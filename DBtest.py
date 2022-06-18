## Bibliothéque du fonctions

from ast import Continue, pattern
from itertools import count
##base du donneés json
from json import load
from multiprocessing.sharedctypes import Value
from pipes import Template
from tkinter.tix import CELL
from types import CellType
## excel
from openpyxl import Workbook, load_workbook, cell
from pyparsing import empty
## base du donneés
from tinydb import TinyDB, Query, where
import json
import re
from EmploiApp import Ui_MainWindow


##creé base du donneés
def DBcreate(): 
 db = TinyDB('db3333.json')
 return db
"""
cmd2='python testbutton.py'
p=subprocess.Popen(cmd2, shell=True)
out, err =p.communicate()
print(err)
print(out)
"""
##excel file
def LoadXLFile(FileName):
    WB = load_workbook(FileName)
    WS = WB.active
    return WS


##Main function cacule max lignes
def searchLastRow(sheet, base):
    for row in range(base, sheet.max_row):
        if sheet.cell(row=row, column=1).value is None: continue
        if next((filter(lambda x: x is not None, [
                sheet.cell(row=i, column=1).value
                for i in range(row + 1, row + 101)
        ])), None) == None:
            return row


##List debut creneaux et fin creneau

## finds all the starts and ends

def parcourirCreneau(WS):
    CoordinatesList = []
    base = 4
    lastrow = searchLastRow(WS, 4)
    row = 4
    table_creneaux = []

    while row <= lastrow:
        creneau = {"debut": WS.cell(row=row, column=1).value, "indice-debut": row}
        row += 1

        while row <= lastrow:
            if WS.cell(row=row, column=1).value is not None:
                break
            row += 1

        end = WS.cell(row=row, column=1).value
        creneau["fin"] = WS.cell(row=row, column=1).value
        creneau["indice-fin"] = row
        table_creneaux.append(creneau)
        CoordinatesList.append(row)
    ##print(creneau)
        row += 1

#list of the sessions
    print(CoordinatesList)

    seances1 = []
    seances2 = []
    seances3 = []
    seances4 = []
    seances5 = []
    seances6 = []
    seances7 = []
    seances8 = []
    seances9 = []
    seances10 = []
    seances11 = []
    seances12 = []
    seances13 = []
    seances14 = []
    seances15 = []
    seances16 = []
    seances17 = []
    seances18 = []
    seances19 = []
    seances20 = []
    seances21 = []
    seances22 = []
    seances23 = []
    seances24 = []
    seances25 = []

##this will become a button in the interface

##Day1
## from 1 - 5 picks the creneaus of the day

##print((CoordinatesList[0]-sizeofcreneau)+counter)
##print(sizeofcreneau)
##print(WS.cell(row=(CoordinatesList[0]-sizeofcreneau)+counter,column=2).value)

    counter = 1
    countercours = 1
    sizeofcreneau = (CoordinatesList[1] + 1) - (CoordinatesList[0] + 1)
## DAY 1
## CRENEAU 1
    if WS.cell(row=(CoordinatesList[0] - sizeofcreneau) + counter,
           column=2).value is not None:
        result = re.search("^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[0] - sizeofcreneau) + counter,
                    column=2).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau1 = WS.iter_rows(min_row=(CoordinatesList[0] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[0],
                            min_col=2,
                            max_col=6)
        CreneauY1 = WS.iter_rows(min_row=(CoordinatesList[0] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[0],
                             min_col=2,
                             max_col=2)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau1:
                    counter += 1
                    seances1.append({
                    'Begins': WS['A4'].value,
                    'Ends': WS.cell(row=CoordinatesList[0], column=1).value,
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:
            for temp in CreneauY1:
            ##print(counter)
                
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances1.append({
                        'Begins': WS['A4'].value,
                        'Ends': WS.cell(row=CoordinatesList[0], column=1).value,
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[0] - sizeofcreneau) + counter,
                           column=2).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[0] - sizeofcreneau) +
                                    counter,
                                    column=2).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[0] - sizeofcreneau) +
                                    counter,
                                    column=2).value)) is None:
                        countercours = 1

    print("Creneau 1 worked")

#CRENAU 2
    counter = 1
    countercours = 1
    if WS.cell(row=(CoordinatesList[1] - sizeofcreneau) + counter,
           column=2).value is not None:
        result = re.search(
        "^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[1] - sizeofcreneau) + counter,
                    column=2).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau2 = WS.iter_rows(min_row=(CoordinatesList[1] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[1],
                            min_col=2,
                            max_col=6)
        CreneauY2 = WS.iter_rows(min_row=(CoordinatesList[1] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[1],
                             min_col=2,
                             max_col=2)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau2:
                    counter += 1
                    seances2.append({
                    'Begins': WS.cell(row=CoordinatesList[0]+1, column=1).value,
                    'Ends': WS.cell(row=CoordinatesList[1], column=1).value,
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:
            for temp in CreneauY2:
                print(counter)
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances2.append({
                        'Begins': WS.cell(row=CoordinatesList[0]+1, column=1).value,
                        'Ends': WS.cell(row=CoordinatesList[1], column=1).value, 
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[1] - sizeofcreneau) + counter,
                           column=2).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[1] - sizeofcreneau) +
                                    counter,
                                    column=2).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[1] - sizeofcreneau) +
                                    counter,
                                    column=2).value)) is None:
                        countercours = 1

    print("Creneau 2 worked")
##CRENNEAU 3
    counter = 1
    countercours = 1
    if WS.cell(row=(CoordinatesList[2] - sizeofcreneau) + counter,
           column=2).value is not None:
        result = re.search(
        "^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[2] - sizeofcreneau) + counter,
                    column=2).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau1 = WS.iter_rows(min_row=(CoordinatesList[2] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[2],
                            min_col=2,
                            max_col=6)
        CreneauY1 = WS.iter_rows(min_row=(CoordinatesList[2] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[2],
                             min_col=2,
                             max_col=2)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau1:
                    counter += 1
                    seances3.append({
                    'Begins': WS.cell(row=CoordinatesList[1]+1, column=1).value,
                    'Ends': WS.cell(row=CoordinatesList[2], column=1).value,
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:

            for temp in CreneauY1:
                print(counter)
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances3.append({
                        'Begins': WS.cell(row=CoordinatesList[1]+1, column=1).value,
                        'Ends': WS.cell(row=CoordinatesList[2], column=1).value, 
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[2] - sizeofcreneau) + counter,
                           column=2).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[2] - sizeofcreneau) +
                                    counter,
                                    column=2).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[2] - sizeofcreneau) +
                                    counter,
                                    column=2).value)) is None:
                        countercours = 1
    print("Creneau 3 worked")
##CRENEAU 4
    counter = 1
    countercours = 1
    if WS.cell(row=(CoordinatesList[3] - sizeofcreneau) + counter,
           column=2).value is not None:
        result = re.search(
        "^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[3] - sizeofcreneau) + counter,
                    column=2).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau1 = WS.iter_rows(min_row=(CoordinatesList[3] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[3],
                            min_col=2,
                            max_col=6)
        CreneauY1 = WS.iter_rows(min_row=(CoordinatesList[3] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[3],
                             min_col=2,
                             max_col=2)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau1:
                    counter += 1
                    seances4.append({
                    'Begins': WS.cell(row=CoordinatesList[2]+1, column=1).value,
                    'Ends': WS.cell(row=CoordinatesList[3], column=1).value,
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:
            for temp in CreneauY1:
                print(counter)
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances4.append({
                        'Begins': WS.cell(row=CoordinatesList[2]+1, column=1).value,
                        'Ends': WS.cell(row=CoordinatesList[3], column=1).value, 
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[3] - sizeofcreneau) + counter,
                           column=2).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[3] - sizeofcreneau) +
                                    counter,
                                    column=2).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[3] - sizeofcreneau) +
                                    counter,
                                    column=2).value)) is None:
                        countercours = 1

    print("Creneau 4 worked")

#CRENEAU 5
    counter = 1
    countercours = 1
    if WS.cell(row=(CoordinatesList[4] - sizeofcreneau) + counter,
           column=2).value is not None:
        result = re.search(
        "^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[4] - sizeofcreneau) + counter,
                    column=2).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau1 = WS.iter_rows(min_row=(CoordinatesList[4] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[4],
                            min_col=2,
                            max_col=6)
        CreneauY1 = WS.iter_rows(min_row=(CoordinatesList[4] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[4],
                             min_col=2,
                             max_col=2)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau1:
                    counter += 1
                    seances5.append({
                    'Begins': WS.cell(row=CoordinatesList[3]+1, column=1).value,
                    'Ends': WS.cell(row=CoordinatesList[4], column=1).value,
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:
            for temp in CreneauY1:
                print(counter)
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances5.append({
                        'Begins': WS.cell(row=CoordinatesList[3]+1, column=1).value,
                        'Ends': WS.cell(row=CoordinatesList[4], column=1).value, 
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[4] - sizeofcreneau) + counter,
                           column=2).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[4] - sizeofcreneau) +
                                    counter,
                                    column=2).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[4] - sizeofcreneau) +
                                    counter,
                                    column=2).value)) is None:
                        countercours = 1

    print("Creneau 5 worked")
## Day 2
##creneau 1
    counter = 1
    countercours = 1
    if WS.cell(row=(CoordinatesList[0] - sizeofcreneau) + counter,
           column=7).value is not None:
        result = re.search(
        "^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[0] - sizeofcreneau) + counter,
                    column=7).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau1 = WS.iter_rows(min_row=(CoordinatesList[0] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[0],
                            min_col=7,
                            max_col=11)
        CreneauY1 = WS.iter_rows(min_row=(CoordinatesList[0] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[0],
                             min_col=7,
                             max_col=7)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau1:
                    counter += 1
                    seances6.append({
                    'Begins': WS['A4'].value,
                    'Ends': WS.cell(row=CoordinatesList[0], column=1).value,
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:
            for temp in CreneauY1:
                print(counter)
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances6.append({
                        'Begins': WS['A4'].value,
                        'Ends': WS.cell(row=CoordinatesList[0], column=1).value, 
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[0] - sizeofcreneau) + counter,
                           column=7).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[0] - sizeofcreneau) +
                                    counter,
                                    column=7).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[0] - sizeofcreneau) +
                                    counter,
                                    column=7).value)) is None:
                        countercours = 1

    print("Creneau 1 day 2 worked")

#CRENAU 2
    counter = 1
    countercours = 1
    if WS.cell(row=(CoordinatesList[1] - sizeofcreneau) + counter,
           column=7).value is not None:
        result = re.search(
        "^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[1] - sizeofcreneau) + counter,
                    column=7).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau1 = WS.iter_rows(min_row=(CoordinatesList[1] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[1],
                            min_col=7,
                            max_col=11)
        CreneauY1 = WS.iter_rows(min_row=(CoordinatesList[1] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[1],
                             min_col=7,
                             max_col=7)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau1:
                    counter += 1
                    seances7.append({
                    'Begins':WS.cell(row=CoordinatesList[0]+1, column=1).value,
                    'Ends': WS.cell(row=CoordinatesList[1], column=1).value,
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:
            for temp in CreneauY1:
                print(counter)
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances7.append({
                        'Begins':WS.cell(row=CoordinatesList[0]+1, column=1).value,
                        'Ends': WS.cell(row=CoordinatesList[1], column=1).value, 
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[1] - sizeofcreneau) + counter,
                           column=7).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[1] - sizeofcreneau) +
                                    counter,
                                    column=7).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[1] - sizeofcreneau) +
                                    counter,
                                    column=7).value)) is None:
                        countercours = 1

    print("Creneau 2 worked")
##CRENNEAU 3
    counter = 1
    countercours = 1
    if WS.cell(row=(CoordinatesList[2] - sizeofcreneau) + counter,
           column=7).value is not None:
        result = re.search(
        "^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[2] - sizeofcreneau) + counter,
                    column=7).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau1 = WS.iter_rows(min_row=(CoordinatesList[2] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[2],
                            min_col=7,
                            max_col=11)
        CreneauY1 = WS.iter_rows(min_row=(CoordinatesList[2] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[2],
                             min_col=7,
                             max_col=7)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau1:
                    counter += 1
                    seances8.append({
                    'Begins':WS.cell(row=CoordinatesList[1]+1, column=1).value,
                    'Ends': WS.cell(row=CoordinatesList[2], column=1).value,
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:
            for temp in CreneauY1:
                print(counter)
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances8.append({
                        'Begins':WS.cell(row=CoordinatesList[1]+1, column=1).value,
                        'Ends': WS.cell(row=CoordinatesList[2], column=1).value, 
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[2] - sizeofcreneau) + counter,
                           column=7).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[2] - sizeofcreneau) +
                                    counter,
                                    column=7).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[2] - sizeofcreneau) +
                                    counter,
                                    column=7).value)) is None:
                        countercours = 1

    print("Creneau 3 worked")
##CRENEAU 4
    counter = 1
    countercours = 1
    if WS.cell(row=(CoordinatesList[3] - sizeofcreneau) + counter,
           column=7).value is not None:
        result = re.search(
        "^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[3] - sizeofcreneau) + counter,
                    column=7).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau1 = WS.iter_rows(min_row=(CoordinatesList[3] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[3],
                            min_col=7,
                            max_col=11)
        CreneauY1 = WS.iter_rows(min_row=(CoordinatesList[3] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[3],
                             min_col=7,
                             max_col=7)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau1:
                    counter += 1
                    seances9.append({
                    'Begins':WS.cell(row=CoordinatesList[2]+1, column=1).value,
                    'Ends': WS.cell(row=CoordinatesList[3], column=1).value,
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:
            for temp in CreneauY1:
                print(counter)
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances9.append({
                        'Begins':WS.cell(row=CoordinatesList[2]+1, column=1).value,
                        'Ends': WS.cell(row=CoordinatesList[3], column=1).value, 
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[3] - sizeofcreneau) + counter,
                           column=7).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[3] - sizeofcreneau) +
                                    counter,
                                    column=7).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[3] - sizeofcreneau) +
                                    counter,
                                    column=7).value)) is None:
                        countercours = 1

    print("Creneau 4 worked")

#CRENEAU 5
    counter = 1
    countercours = 1
    if WS.cell(row=(CoordinatesList[4] - sizeofcreneau) + counter,
           column=7).value is not None:
        result = re.search(
        "^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[4] - sizeofcreneau) + counter,
                    column=7).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau1 = WS.iter_rows(min_row=(CoordinatesList[4] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[4],
                            min_col=7,
                            max_col=11)
        CreneauY1 = WS.iter_rows(min_row=(CoordinatesList[4] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[4],
                             min_col=7,
                             max_col=7)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau1:
                    counter += 1
                    seances10.append({
                    'Begins':WS.cell(row=CoordinatesList[3]+1, column=1).value,
                    'Ends': WS.cell(row=CoordinatesList[4], column=1).value,
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:
            for temp in CreneauY1:
                print(counter)
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances10.append({
                        'Begins':WS.cell(row=CoordinatesList[3]+1, column=1).value,
                        'Ends': WS.cell(row=CoordinatesList[4], column=1).value, 
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[4] - sizeofcreneau) + counter,
                           column=7).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[4] - sizeofcreneau) +
                                    counter,
                                    column=7).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[4] - sizeofcreneau) +
                                    counter,
                                    column=7).value)) is None:
                        countercours = 1

    print("Creneau 5 worked")
## Day 3
##creneau 1
    counter = 1
    countercours = 1
    if WS.cell(row=(CoordinatesList[0] - sizeofcreneau) + counter,
           column=12).value is not None:
        result = re.search(
        "^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[0] - sizeofcreneau) + counter,
                    column=12).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau1 = WS.iter_rows(min_row=(CoordinatesList[0] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[0],
                            min_col=12,
                            max_col=16)
        CreneauY1 = WS.iter_rows(min_row=(CoordinatesList[0] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[0],
                             min_col=12,
                             max_col=12)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau1:
                    counter += 1
                    seances11.append({
                    'Begins': WS['A4'].value,
                    'Ends': WS.cell(row=CoordinatesList[0], column=1).value,
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:
            for temp in CreneauY1:
                print(counter)
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances11.append({
                        'Begins': WS['A4'].value,
                        'Ends': WS.cell(row=CoordinatesList[0], column=1).value, 
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[0] - sizeofcreneau) + counter,
                           column=12).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[0] - sizeofcreneau) +
                                    counter,
                                    column=12).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[0] - sizeofcreneau) +
                                    counter,
                                    column=12).value)) is None:
                        countercours = 1

#CRENAU 2
    counter = 1
    countercours = 1
    if WS.cell(row=(CoordinatesList[1] - sizeofcreneau) + counter,
           column=12).value is not None:
        result = re.search(
        "^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[1] - sizeofcreneau) + counter,
                    column=12).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau1 = WS.iter_rows(min_row=(CoordinatesList[1] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[1],
                            min_col=12,
                            max_col=16)
        CreneauY1 = WS.iter_rows(min_row=(CoordinatesList[1] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[1],
                             min_col=12,
                             max_col=12)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau1:
                    counter += 1
                    seances12.append({
                    'Begins': WS.cell(row=CoordinatesList[0]+1, column=1).value,
                    'Ends': WS.cell(row=CoordinatesList[1], column=1).value,
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:
            for temp in CreneauY1:
                print(counter)
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances12.append({
                        'Begins': WS.cell(row=CoordinatesList[0]+1, column=1).value,
                        'Ends': WS.cell(row=CoordinatesList[1], column=1).value, 
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[1] - sizeofcreneau) + counter,
                           column=12).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[1] - sizeofcreneau) +
                                    counter,
                                    column=12).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[1] - sizeofcreneau) +
                                    counter,
                                    column=12).value)) is None:
                        countercours = 1

##CRENNEAU 3
    counter = 1
    countercours = 1
    if WS.cell(row=(CoordinatesList[2] - sizeofcreneau) + counter,
           column=12).value is not None:
        result = re.search(
        "^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[2] - sizeofcreneau) + counter,
                    column=12).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau1 = WS.iter_rows(min_row=(CoordinatesList[2] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[2],
                            min_col=12,
                            max_col=16)
        CreneauY1 = WS.iter_rows(min_row=(CoordinatesList[2] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[2],
                             min_col=12,
                             max_col=12)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau1:
                    counter += 1
                    seances13.append({
                    'Begins': WS.cell(row=CoordinatesList[1]+1, column=1).value,
                    'Ends': WS.cell(row=CoordinatesList[2], column=1).value,
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:
            for temp in CreneauY1:
                print(counter)
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances13.append({
                        'Begins': WS.cell(row=CoordinatesList[1]+1, column=1).value,
                        'Ends': WS.cell(row=CoordinatesList[2], column=1).value, 
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[2] - sizeofcreneau) + counter,
                           column=12).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[2] - sizeofcreneau) +
                                    counter,
                                    column=12).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[2] - sizeofcreneau) +
                                    counter,
                                    column=12).value)) is None:
                        countercours = 1

##CRENEAU 4
    counter = 1
    countercours = 1
    if WS.cell(row=(CoordinatesList[3] - sizeofcreneau) + counter,
           column=12).value is not None:
        result = re.search(
        "^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[3] - sizeofcreneau) + counter,
                    column=12).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau1 = WS.iter_rows(min_row=(CoordinatesList[3] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[3],
                            min_col=12,
                            max_col=16)
        CreneauY1 = WS.iter_rows(min_row=(CoordinatesList[3] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[3],
                             min_col=12,
                             max_col=12)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau1:
                    counter += 1
                    seances14.append({
                    'Begins': WS.cell(row=CoordinatesList[2]+1, column=1).value,
                    'Ends': WS.cell(row=CoordinatesList[3], column=1).value,
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:
            for temp in CreneauY1:
                print(counter)
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances14.append({
                        'Begins': WS.cell(row=CoordinatesList[2]+1, column=1).value,
                        'Ends': WS.cell(row=CoordinatesList[3], column=1).value, 
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[3] - sizeofcreneau) + counter,
                           column=12).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[3] - sizeofcreneau) +
                                    counter,
                                    column=12).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[3] - sizeofcreneau) +
                                    counter,
                                    column=12).value)) is None:
                        countercours = 1

#CRENEAU 5
    counter = 1
    countercours = 1
    if WS.cell(row=(CoordinatesList[4] - sizeofcreneau) + counter,
           column=12).value is not None:
        result = re.search(
        "^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[4] - sizeofcreneau) + counter,
                    column=12).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau1 = WS.iter_rows(min_row=(CoordinatesList[4] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[4],
                            min_col=12,
                            max_col=16)
        CreneauY1 = WS.iter_rows(min_row=(CoordinatesList[4] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[4],
                             min_col=12,
                             max_col=12)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau1:
                    counter += 1
                    seances15.append({
                    'Begins': WS.cell(row=CoordinatesList[3]+1, column=1).value,
                    'Ends': WS.cell(row=CoordinatesList[4], column=1).value,
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:
            for temp in CreneauY1:
                print(counter)
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances15.append({
                        'Begins': WS.cell(row=CoordinatesList[3]+1, column=1).value,
                        'Ends': WS.cell(row=CoordinatesList[4], column=1).value, 
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[4] - sizeofcreneau) + counter,
                           column=12).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[4] - sizeofcreneau) +
                                    counter,
                                    column=12).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[4] - sizeofcreneau) +
                                    counter,
                                    column=12).value)) is None:
                        countercours = 1

## Day 4
##creneau 1
    counter = 1
    countercours = 1
    if WS.cell(row=(CoordinatesList[0] - sizeofcreneau) + counter,
           column=17).value is not None:
        result = re.search(
        "^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[0] - sizeofcreneau) + counter,
                    column=17).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau1 = WS.iter_rows(min_row=(CoordinatesList[0] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[0],
                            min_col=17,
                            max_col=21)
        CreneauY1 = WS.iter_rows(min_row=(CoordinatesList[0] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[0],
                             min_col=17,
                             max_col=17)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau1:
                    counter += 1
                    seances16.append({
                    'Begins': WS['A4'].value,
                    'Ends': WS.cell(row=CoordinatesList[0], column=1).value,
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:
            for temp in CreneauY1:
                print(counter)
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances16.append({
                        'Begins': WS['A4'].value,
                        'Ends': WS.cell(row=CoordinatesList[0], column=1).value, 
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[0] - sizeofcreneau) + counter,
                           column=17).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[0] - sizeofcreneau) +
                                    counter,
                                    column=17).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[0] - sizeofcreneau) +
                                    counter,
                                    column=17).value)) is None:
                        countercours = 1

#CRENAU 2
    counter = 1
    countercours = 1
    if WS.cell(row=(CoordinatesList[1] - sizeofcreneau) + counter,
           column=17).value is not None:
        result = re.search(
        "^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[1] - sizeofcreneau) + counter,
                    column=17).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau1 = WS.iter_rows(min_row=(CoordinatesList[1] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[1],
                            min_col=17,
                            max_col=21)
        CreneauY1 = WS.iter_rows(min_row=(CoordinatesList[1] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[1],
                             min_col=17,
                             max_col=17)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau1:
                    counter += 1
                    seances17.append({
                    'Begins': WS.cell(row=CoordinatesList[0]+1, column=1).value,
                    'Ends': WS.cell(row=CoordinatesList[1], column=1).value,
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:
            for temp in CreneauY1:
                print(counter)
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances17.append({
                        'Begins': WS.cell(row=CoordinatesList[0]+1, column=1).value,
                        'Ends': WS.cell(row=CoordinatesList[1], column=1).value, 
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[1] - sizeofcreneau) + counter,
                           column=17).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[1] - sizeofcreneau) +
                                    counter,
                                    column=17).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[1] - sizeofcreneau) +
                                    counter,
                                    column=17).value)) is None:
                        countercours = 1

##CRENNEAU 3
    counter = 1
    countercours = 1
    if WS.cell(row=(CoordinatesList[2] - sizeofcreneau) + counter,
           column=17).value is not None:
        result = re.search(
        "^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[2] - sizeofcreneau) + counter,
                    column=17).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau1 = WS.iter_rows(min_row=(CoordinatesList[2] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[2],
                            min_col=17,
                            max_col=21)
        CreneauY1 = WS.iter_rows(min_row=(CoordinatesList[2] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[2],
                             min_col=17,
                             max_col=17)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau1:
                    counter += 1
                    seances18.append({
                    'Begins': WS.cell(row=CoordinatesList[1]+1, column=1).value,
                    'Ends': WS.cell(row=CoordinatesList[2], column=1).value,
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:
            for temp in CreneauY1:
                print(counter)
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances18.append({
                        'Begins': WS.cell(row=CoordinatesList[1]+1, column=1).value,
                        'Ends': WS.cell(row=CoordinatesList[2], column=1).value, 
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[2] - sizeofcreneau) + counter,
                           column=17).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[2] - sizeofcreneau) +
                                    counter,
                                    column=17).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[2] - sizeofcreneau) +
                                    counter,
                                    column=17).value)) is None:
                        countercours = 1

##CRENEAU 4
    counter = 1
    countercours = 1
    if WS.cell(row=(CoordinatesList[3] - sizeofcreneau) + counter,
           column=17).value is not None:
        result = re.search(
        "^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[3] - sizeofcreneau) + counter,
                    column=17).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau1 = WS.iter_rows(min_row=(CoordinatesList[3] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[3],
                            min_col=17,
                            max_col=21)
        CreneauY1 = WS.iter_rows(min_row=(CoordinatesList[3] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[3],
                             min_col=17,
                             max_col=17)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau1:
                    counter += 1
                    seances19.append({
                    'Begins': WS.cell(row=CoordinatesList[2]+1, column=1).value,
                    'Ends': WS.cell(row=CoordinatesList[3], column=1).value,
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:
            for temp in CreneauY1:
                print(counter)
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances19.append({
                        'Begins': WS.cell(row=CoordinatesList[2]+1, column=1).value,
                        'Ends': WS.cell(row=CoordinatesList[3], column=1).value, 
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[3] - sizeofcreneau) + counter,
                           column=17).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[3] - sizeofcreneau) +
                                    counter,
                                    column=17).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[3] - sizeofcreneau) +
                                    counter,
                                    column=17).value)) is None:
                        countercours = 1

#CRENEAU 5
    counter = 1
    countercours = 1
    if WS.cell(row=(CoordinatesList[4] - sizeofcreneau) + counter,
           column=17).value is not None:
        result = re.search(
        "^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[4] - sizeofcreneau) + counter,
                    column=17).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau1 = WS.iter_rows(min_row=(CoordinatesList[4] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[4],
                            min_col=17,
                            max_col=21)
        CreneauY1 = WS.iter_rows(min_row=(CoordinatesList[4] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[4],
                             min_col=17,
                             max_col=17)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau1:
                    counter += 1
                    seances20.append({
                    'Begins': WS.cell(row=CoordinatesList[3]+1, column=1).value,
                    'Ends': WS.cell(row=CoordinatesList[4], column=1).value,
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:
            for temp in CreneauY1:
                print(counter)
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances20.append({
                        'Begins': WS.cell(row=CoordinatesList[3]+1, column=1).value,
                        'Ends': WS.cell(row=CoordinatesList[4], column=1).value, 
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[4] - sizeofcreneau) + counter,
                           column=17).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[4] - sizeofcreneau) +
                                    counter,
                                    column=17).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[4] - sizeofcreneau) +
                                    counter,
                                    column=17).value)) is None:
                        countercours = 1

## Day 5
##creneau 1
    counter = 1
    countercours = 1
    if WS.cell(row=(CoordinatesList[0] - sizeofcreneau) + counter,
           column=22).value is not None:
        result = re.search(
        "^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[0] - sizeofcreneau) + counter,
                    column=22).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau1 = WS.iter_rows(min_row=(CoordinatesList[0] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[0],
                            min_col=22,
                            max_col=26)
        CreneauY1 = WS.iter_rows(min_row=(CoordinatesList[0] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[0],
                             min_col=22,
                             max_col=22)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau1:
                    counter += 1
                    seances21.append({
                    'Begins': WS['A4'].value,
                    'Ends': WS.cell(row=CoordinatesList[0], column=1).value,
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:
            for temp in CreneauY1:
                print(counter)
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances21.append({
                        'Begins': WS['A4'].value,
                        'Ends': WS.cell(row=CoordinatesList[0], column=1).value, 
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[0] - sizeofcreneau) + counter,
                           column=22).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[0] - sizeofcreneau) +
                                    counter,
                                    column=22).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[0] - sizeofcreneau) +
                                    counter,
                                    column=22).value)) is None:
                        countercours = 1

#CRENAU 2
    counter = 1
    countercours = 1
    if WS.cell(row=(CoordinatesList[1] - sizeofcreneau) + counter,
           column=22).value is not None:
        result = re.search(
        "^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[1] - sizeofcreneau) + counter,
                    column=22).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau1 = WS.iter_rows(min_row=(CoordinatesList[1] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[1],
                            min_col=22,
                            max_col=26)
        CreneauY1 = WS.iter_rows(min_row=(CoordinatesList[1] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[1],
                             min_col=22,
                             max_col=22)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau1:
                    counter += 1
                    seances22.append({
                    'Begins': WS.cell(row=CoordinatesList[0]+1, column=1).value,
                    'Ends': WS.cell(row=CoordinatesList[1], column=1).value,
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:
            for temp in CreneauY1:
                print(counter)
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances22.append({
                        'Begins': WS.cell(row=CoordinatesList[0]+1, column=1).value,
                        'Ends': WS.cell(row=CoordinatesList[1], column=1).value, 
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[1] - sizeofcreneau) + counter,
                           column=22).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[1] - sizeofcreneau) +
                                    counter,
                                    column=22).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[1] - sizeofcreneau) +
                                    counter,
                                    column=22).value)) is None:
                        countercours = 1

##CRENNEAU 3
    counter = 1
    countercours = 1
    if WS.cell(row=(CoordinatesList[2] - sizeofcreneau) + counter,
           column=22).value is not None:
        result = re.search(
        "^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[2] - sizeofcreneau) + counter,
                    column=22).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau1 = WS.iter_rows(min_row=(CoordinatesList[2] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[2],
                            min_col=22,
                            max_col=26)
        CreneauY1 = WS.iter_rows(min_row=(CoordinatesList[2] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[2],
                             min_col=22,
                             max_col=22)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau1:
                    counter += 1
                    seances23.append({
                    'Begins': WS.cell(row=CoordinatesList[1]+1, column=1).value,
                    'Ends': WS.cell(row=CoordinatesList[2], column=1).value,
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:
            for temp in CreneauY1:
                print(counter)
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances23.append({
                        'Begins': WS.cell(row=CoordinatesList[1]+1, column=1).value,
                        'Ends': WS.cell(row=CoordinatesList[2], column=1).value, 
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[2] - sizeofcreneau) + counter,
                           column=22).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[2] - sizeofcreneau) +
                                    counter,
                                    column=22).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[2] - sizeofcreneau) +
                                    counter,
                                    column=22).value)) is None:
                        countercours = 1

##CRENEAU 4
    counter = 1
    countercours = 1
    if WS.cell(row=(CoordinatesList[3] - sizeofcreneau) + counter,
           column=22).value is not None:
        result = re.search(
        "^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[3] - sizeofcreneau) + counter,
                    column=22).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau1 = WS.iter_rows(min_row=(CoordinatesList[3] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[3],
                            min_col=22,
                            max_col=26)
        CreneauY1 = WS.iter_rows(min_row=(CoordinatesList[3] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[3],
                             min_col=22,
                             max_col=22)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau1:
                    counter += 1
                    seances24.append({
                    'Begins': WS.cell(row=CoordinatesList[2]+1, column=1).value,
                    'Ends': WS.cell(row=CoordinatesList[3], column=1).value,
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:
            for temp in CreneauY1:
                print(counter)
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances24.append({
                        'Begins': WS.cell(row=CoordinatesList[2]+1, column=1).value,
                        'Ends': WS.cell(row=CoordinatesList[3], column=1).value, 
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[3] - sizeofcreneau) + counter,
                           column=22).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[3] - sizeofcreneau) +
                                    counter,
                                    column=22).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[3] - sizeofcreneau) +
                                    counter,
                                    column=22).value)) is None:
                        countercours = 1

#CRENEAU 5
    counter = 1
    countercours = 1
    if WS.cell(row=(CoordinatesList[4] - sizeofcreneau) + counter,
           column=22).value is not None:
        result = re.search(
        "^G[1-9]$",
        str(
            WS.cell(row=(CoordinatesList[4] - sizeofcreneau) + counter,
                    column=22).value))
    else:
        result = None

    while counter < sizeofcreneau:
        Creneau1 = WS.iter_rows(min_row=(CoordinatesList[4] - sizeofcreneau) +
                            counter,
                            max_row=CoordinatesList[4],
                            min_col=22,
                            max_col=26)
        CreneauY1 = WS.iter_rows(min_row=(CoordinatesList[4] - sizeofcreneau) +
                             counter,
                             max_row=CoordinatesList[4],
                             min_col=22,
                             max_col=22)
        if result != None:
        ##print((CoordinatesList[0]-sizeofcreneau)+counter)
            while result != None and counter < sizeofcreneau:
                for Group, TypeSceance, Sceance, ProfName, Salle in Creneau1:
                    counter += 1
                    seances25.append({
                    'Begins': WS.cell(row=CoordinatesList[3]+1, column=1).value,
                    'Ends': WS.cell(row=CoordinatesList[4], column=1).value, 
                    'Group': Group.value,
                    'TypeSceance': TypeSceance.value,
                    'sceance': Sceance.value,
                    'ProfName': ProfName.value,
                    'Salle': Salle.value,
                })

        else:
            for temp in CreneauY1:
                print(counter)
                if countercours == 1:
                    Value1=temp[0].value
                    counter += 1
                    countercours += 1
                elif countercours == 2:
                    Value2=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 3:
                    Value3=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 4:
                    Value4=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 5:
                    Value5=temp[0].value
                    counter += 1
                    countercours += 1

                elif countercours == 6:
                    seances25.append({
                        'Begins': WS.cell(row=CoordinatesList[3]+1, column=1).value,
                        'Ends': WS.cell(row=CoordinatesList[4], column=1).value, 
                        'CasVide': Value1,
                        'NomMod': Value2,
                        'Prof': Value3,
                        'Salle_Cour': Value4,
                        'TypeSC': Value5,
                        'CasVide2': temp[0].value})
                    counter += 1
                    countercours += 1
                elif counter > 6 and temp[0].value == None:
                    counter += 1
                    countercours += 1
                    if WS.cell(row=(CoordinatesList[4] - sizeofcreneau) + counter,
                           column=22).value is not None:
                        result = re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[4] - sizeofcreneau) +
                                    counter,
                                    column=22).value))
                    elif re.search(
                        "^G[1-9]$",
                        str(
                            WS.cell(row=(CoordinatesList[4] - sizeofcreneau) +
                                    counter,
                                    column=22).value)) is None:
                        countercours = 1
    return CoordinatesList,seances1,seances2,seances3,seances4,seances5,seances6,seances7,seances8,seances9,seances10,seances11,seances12,seances13,seances14,seances15,seances16,seances17,seances18,seances19,seances20,seances21,seances22,seances23,seances24,seances25



##inserts in the data base

##formation= input


def FillData(db,WS,FormationInput,NiveauInput,SemestreInput,AnnéeInput,CoordinatesList, seances1, seances2, seances3, seances4, seances5, seances6, seances7, seances8, seances9, seances10, seances11, seances12, seances13, seances14, seances15, seances16, seances17, seances18, seances19, seances20, seances21, seances22, seances23, seances24, seances25):
 db.insert({
        'formation':FormationInput,
        'Niveau': NiveauInput,
        'Semestre': SemestreInput,
        'Année':AnnéeInput,
        'Week': [{
            "dimanche" : [{
                'start': WS['A4'].value,
                'end': WS.cell(row=CoordinatesList[0], column=1).value,
                'sceances': seances1,
            }, {
                'start':
                WS.cell(row=CoordinatesList[0] + 1, column=1).value,
                'end':
                WS.cell(row=CoordinatesList[1], column=1).value,
                'sceances': seances2,
            }, {
                'start':
                WS.cell(row=CoordinatesList[1] + 1, column=1).value,
                'end':
                WS.cell(row=CoordinatesList[2], column=1).value,
                'sceances':
                seances3,
            }, {
                'start':
                WS.cell(row=CoordinatesList[2] + 1, column=1).value,
                'end':
                WS.cell(row=CoordinatesList[3], column=1).value,
                'sceances':
                seances4,
            }, {
                'start':
                WS.cell(row=CoordinatesList[3] + 1, column=1).value,
                'end':
                WS.cell(row=CoordinatesList[4], column=1).value,
                'sceances':
                seances5,
            }],
            'Lundi': [{
                'start': WS['A4'].value,
                'end': WS.cell(row=CoordinatesList[0], column=1).value,
                'sceances': seances6,
            }, {
                'start':
                WS.cell(row=CoordinatesList[0] + 1, column=1).value,
                'end':
                WS.cell(row=CoordinatesList[1], column=1).value,
                'sceances':
                seances7,
            }, {
                'start':
                WS.cell(row=CoordinatesList[1] + 1, column=1).value,
                'end':
                WS.cell(row=CoordinatesList[2], column=1).value,
                'sceances':
                seances8,
            }, {
                'start':
                WS.cell(row=CoordinatesList[2] + 1, column=1).value,
                'end':
                WS.cell(row=CoordinatesList[3], column=1).value,
                'sceances':
                seances9,
            }, {
                'start':
                WS.cell(row=CoordinatesList[3] + 1, column=1).value,
                'end':
                WS.cell(row=CoordinatesList[4], column=1).value,
                'sceances':
                seances10,
            }],
            'Mardi': [{
                'start': WS['A4'].value,
                'end': WS.cell(row=CoordinatesList[0], column=1).value,
                'sceances': seances11,
            }, {
                'start':
                WS.cell(row=CoordinatesList[0] + 1, column=1).value,
                'end':
                WS.cell(row=CoordinatesList[1], column=1).value,
                'sceances':
                seances12,
            }, {
                'start':
                WS.cell(row=CoordinatesList[1] + 1, column=1).value,
                'end':
                WS.cell(row=CoordinatesList[2], column=1).value,
                'sceances':
                seances13,
            }, {
                'start':
                WS.cell(row=CoordinatesList[2] + 1, column=1).value,
                'end':
                WS.cell(row=CoordinatesList[3], column=1).value,
                'sceances':
                seances14,
            }, {
                'start':
                WS.cell(row=CoordinatesList[3] + 1, column=1).value,
                'end':
                WS.cell(row=CoordinatesList[4], column=1).value,
                'sceances':
                seances15,
            }],
            'Mercredi': [{
                'start': WS['A4'].value,
                'end': WS.cell(row=CoordinatesList[0], column=1).value,
                'sceances': seances16,
            }, {
                'start':
                WS.cell(row=CoordinatesList[0] + 1, column=1).value,
                'end':
                WS.cell(row=CoordinatesList[1], column=1).value,
                'sceances':
                seances17,
            }, {
                'start':
                WS.cell(row=CoordinatesList[1] + 1, column=1).value,
                'end':
                WS.cell(row=CoordinatesList[2], column=1).value,
                'sceances':
                seances18,
            }, {
                'start':
                WS.cell(row=CoordinatesList[2] + 1, column=1).value,
                'end':
                WS.cell(row=CoordinatesList[3], column=1).value,
                'sceances':
                seances19,
            }, {
                'start':
                WS.cell(row=CoordinatesList[3] + 1, column=1).value,
                'end':
                WS.cell(row=CoordinatesList[4], column=1).value,
                'sceances':
                seances20,
            }],
            'Jeudi': [{
                'start': WS['A4'].value,
                'end': WS.cell(row=CoordinatesList[0], column=1).value,
                'sceances': seances21,
            }, {
                'start':
                WS.cell(row=CoordinatesList[0] + 1, column=1).value,
                'end':
                WS.cell(row=CoordinatesList[1], column=1).value,
                'sceances':
                seances22,
            }, {
                'start':
                WS.cell(row=CoordinatesList[1] + 1, column=1).value,
                'end':
                WS.cell(row=CoordinatesList[2], column=1).value,
                'sceances':
                seances23,
            }, {
                'start':
                WS.cell(row=CoordinatesList[2] + 1, column=1).value,
                'end':
                WS.cell(row=CoordinatesList[3], column=1).value,
                'sceances':
                seances24,
            }, {
                'start':
                WS.cell(row=CoordinatesList[3] + 1, column=1).value,
                'end':
                WS.cell(row=CoordinatesList[4], column=1).value,
                'sceances':
                seances25,
            }],
        }]
    })








def printData(db):
 for item in db:
    print(item)

 db.all()





##Try fixing req3 4 6
##db.search(Req1.Week.any(Req2.dimanche.any(Query().start == "08H00")))
##db.search(Req1.Week.any(Req2.dimanche.any(Req3.Lundi.any(Req4.Mardi.any(Req5.Mercredi.any(Req6.Jeudi.any(Query().start == "08H00")))))))
def ResultFinder(db):
 Req1 = Query()
 Req2 = Query()
 Req3 = Query()
 Req4 = Query()
 Req5 = Query()
 Req6 = Query() 
 Result = db.search(Req1.Week.any(Req2.dimanche.any(Query().start == "08H00")))
 return Result
#(User.Week.Creneaux.start ==  '08H00')


def testField(value, test):
    return True if test == "" else value == test


def Search(Result,niveau, formation, semestre, Année, group, prof, salle):
    res1 = []
    res2 = []
    res3 = []
    res4 = []
    res5 = []
    StartsAndEnds = []
    ListFormation = []
    ListNiveau = []
    ListSemestre = []
    ListYear = []
    cpt2=0
    testing = group
    counter=['dimanche','Lundi','Mardi','Mercredi','Jeudi']
    for obj in Result:
     
      

     
     if testField(obj["formation"], formation) and testField(obj["Niveau"], niveau) and testField(obj["Semestre"], semestre): ##and testField(obj["Année"], Année):  #rajouter les tests sur l'année et le semestre
            for j in obj["Week"]:
                for cpt in range(0, 5):
                    
                    for cr in j[counter[cpt]]:
                     StartsAndEnds.append(cr["start"])
                     StartsAndEnds.append(cr["end"])
                     cpt2=0
                     for s in cr["sceances"]:
                      
                      if "Group" in s:
                        if testField(s["Group"], group) and testField(s["ProfName"], prof) and testField(s["Salle"], salle):
                            if cpt==0:
                             res1.append(s)
                             print(s)
                            elif cpt==1:
                             res2.append(s)
                             print(s)
                            elif cpt==2:
                             res3.append(s)
                             print(s)
                            elif cpt==3:
                             res4.append(s)
                             print(s)
                            else:
                             res5.append(s)
                             print(s)
                      elif "CasVide" in s:
                            if testing.startswith('G') and cpt == 0:
                                res1.append(s)
                                print(s)
                            elif testing.startswith('G') and cpt==1:
                                res2.append(s)
                                print(s)
                            elif testing.startswith('G') and cpt==2:
                                res3.append(s)
                                print(s)
                            elif testing.startswith('G') and cpt==3:
                                res4.append(s)
                                print(s)
                            elif testing.startswith('G') and cpt==4:
                                res5.append(s)
                                print(s)
                            elif testField(s["Prof"], prof) and testField(s["Salle_Cour"], salle):
                                if cpt==0:
                                    res1.append(s)
                                    print(s)
                                elif cpt==1:
                                    res2.append(s)
                                    print(s)
                                elif cpt==2:
                                    res3.append(s)
                                    print(s)
                                elif cpt==3:
                                    res4.append(s)
                                    print(s)
                                else:
                                    res5.append(s)
                                    print(s)

                         
                      
                      
                      """
                        if "Prof" in s:                        ##testField(s["Prof"], prof) and testField(s["Salle_Cour"], salle) and group.startswith('G')==True:
                            if testField(s["Prof"], prof):
                             res1.append(s)
                        elif "Salle_Cour" in s:
                            if testField(s["Salle_Cour"], salle):
                             res1.append(s) 
                        elif group.startswith('G')==True:
                            print("cours worked worked")
                            res1.append(s)
                          """
     
     
     
     return res1,res2,res3,res4,res5,StartsAndEnds

def EmptyDB(db):
 db.truncate()
"""
for cpt in range(0, 4):
        for item in obj:
            if cpt == 0:
                ListFormation.append(item)
            elif cpt == 1:
                ListNiveau.append(item)
            elif cpt == 2:
                ListSemestre.append(item)
            elif cpt==3:
                ListYear.append(item)
"""